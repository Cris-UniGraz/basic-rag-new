import os
import requests
import argparse
import time
import json
from pathlib import Path
from tqdm import tqdm

# PARA USAR ESTE SCRIPT:
# python load_documents.py --dir "C:/Pruebas/RAG Search/demo_docu_4_min/" --collection uni_test_1_1
# python load_documents.py --dir "C:/Pruebas/RAG Search/documentos_idioma_all/" --collection uni_docs_1_0
# python load_documents.py --url http://143.50.27.65:8000 --dir "C:/Pruebas/RAG Search/demo_docu_4_min/" --collection uni_test_2_0
# python load_documents.py --url http://143.50.27.65:8000 --dir "C:/Pruebas/RAG Search/documentos_idioma_all/" --collection uni_docs_1_0
#
# DEPENDENCIAS REQUERIDAS:
# pip install requests tqdm PyMuPDF docx2txt openpyxl

def get_files_recursively(directory):
    """Obtiene todos los archivos de un directorio y sus subdirectorios."""
    files = []
    for root, _, filenames in os.walk(directory):
        for filename in filenames:
            if not filename.startswith('~') and not filename.startswith('.'):
                file_path = os.path.join(root, filename)
                files.append(file_path)
    return files

def upload_documents(base_url, directory_path, collection_name):
    """Carga documentos al servicio RAG desde subcarpetas por idioma."""
    # Verificar que el directorio exista
    if not os.path.exists(directory_path):
        print(f"Error: El directorio {directory_path} no existe.")
        return False

    # Verificar las subcarpetas de idioma
    de_dir = os.path.join(directory_path, "de")
    en_dir = os.path.join(directory_path, "en")

    if not (os.path.exists(de_dir) or os.path.exists(en_dir)):
        print(f"Error: No se encontraron las subcarpetas 'de' o 'en' en {directory_path}.")
        return False

    print(f"\n📄 Utilidad de carga masiva de documentos para RAG")
    print(f"==================================================")
    print(f"URL Base: {base_url}")
    print(f"Directorio: {directory_path}")
    print(f"Nombre base de colección: {collection_name}")

    success = True

    # Procesar documentos en alemán
    if os.path.exists(de_dir):
        de_success = process_language_directory(base_url, de_dir, "german", collection_name)
        success = success and de_success

    # Procesar documentos en inglés
    if os.path.exists(en_dir):
        en_success = process_language_directory(base_url, en_dir, "english", collection_name)
        success = success and en_success
        
    # En cualquier caso, consideramos que el proceso fue exitoso mientras estemos cargando archivos

    if success:
        print(f"\n✅ Procesamiento de todos los documentos completado con éxito!")
    else:
        print(f"\n⚠️ Procesamiento completado con advertencias. Los documentos se cargaron, pero puede haber algún problema de configuración en la colección.")
    
    return success


def process_language_directory(base_url, directory, language, collection_name):
    """Procesa documentos de una carpeta de idioma específica."""
    # Obtener todos los archivos del directorio y sus subdirectorios
    files = get_files_recursively(directory)

    if not files:
        print(f"No se encontraron archivos en {directory}.")
        return True

    lang_name = "alemán" if language == "german" else "inglés"
    print(f"\nProcesando {len(files)} documentos en {lang_name} desde {directory}...")

    # Mostrar barra de progreso para el procesamiento total
    file_progress = tqdm(total=len(files), desc=f"Progreso general", unit="archivo")
    
    # Variables para seguimiento de éxitos y errores
    successful_files = 0
    failed_files = []

    # Procesar cada archivo individualmente
    
    for file_path in files:
        file_name = os.path.basename(file_path)
        
        try:
            # Determinar número de páginas, dimensiones y metadatos según tipo, asegurando valores por defecto
            total_pages = get_document_pages(file_path) or 1  # Si retorna None o 0, usar 1
            width, height = get_document_dimensions(file_path)  # Obtener ancho y alto
            width = width or 612  # Si retorna None o 0, usar 612
            height = height or 864  # Si retorna None o 0, usar 864
            sheet_index = get_sheet_index(file_path)  # Por defecto es 0
            sheet_name = get_sheet_name(file_path)  # Por defecto es vacío
            
            # Depuración para ver qué valores se están calculando
            print(f"\nDebug - Archivo: {file_name}")
            print(f"Debug - Extensión: {Path(file_path).suffix.lower()}")
            print(f"Debug - Páginas calculadas: {total_pages}")
            print(f"Debug - Dimensiones calculadas: {width}x{height}")
            print(f"Debug - Sheet index: {sheet_index}")
            print(f"Debug - Sheet name: {sheet_name}")
            
            # Preparar archivo para subir
            files_to_upload = [
                ('files', (file_name, open(file_path, 'rb'), 'application/octet-stream'))
            ]
            
            # Preparar datos específicos según el tipo de documento
            ext = Path(file_path).suffix.lower().lstrip('.')
            
            # Enfoque drásticamente simplificado: conjunto mínimo de campos comunes
            # Basado exactamente en lo que hace local-adv-rag, pero con campos adicionales requeridos
            
            # Lista completa de todos los campos conocidos requeridos por Milvus
            all_required_fields = {
                # Campos administrativos
                'language': language,
                'collection_name': collection_name,
                
                # Campos de metadatos comunes
                'source': file_name,
                'file_type': 'text',          # Valor por defecto
                'page_number': '1',           # Valor por defecto
                'sheet_name': '',             # Valor por defecto
                'sheet_index': '0',           # Valor por defecto
                
                # Campos específicos de dimensiones y paginación
                'total_pages': str(int(total_pages)),
                'total_sheets': '1',           # Valor por defecto
                'width': str(int(width)),
                'height': str(int(height))
            }
            
            # Parámetros base para la API
            upload_data = all_required_fields.copy()
            
            # Sobrescribir campos específicos según el tipo de archivo
            if ext in ['xls', 'xlsx']:
                # Excel - valores específicos
                upload_data.update({
                    'file_type': 'excel',       # Tipo fijo para simplificar
                    'page_number': '-1',        # Excel usa -1 en local-adv-rag
                    'sheet_name': sheet_name,
                    'total_sheets': str(int(total_pages))
                })
            elif ext in ['doc', 'docx']:
                # Word - valores específicos
                upload_data.update({
                    'file_type': 'word',        # Tipo fijo para simplificar
                })
            elif ext == 'pdf':
                # PDF - valores específicos
                upload_data.update({
                    'file_type': 'pdf',         # Tipo fijo
                })
            
            # Depurar los datos que se enviarán
            print(f"Debug - Enviando a la API: {upload_data}")
            
            # Subir documento
            try:
                # Ajustar timeout basado en tamaño del archivo
                file_size = os.path.getsize(file_path)
                # Si el archivo es grande (>5MB), dar más tiempo para carga
                upload_timeout = 120 if file_size > 5*1024*1024 else 60
                
                response = requests.post(
                    f"{base_url}/api/documents/upload",
                    files=files_to_upload,
                    data=upload_data,
                    timeout=upload_timeout
                )

                response.raise_for_status()
                result = response.json()

                # Monitorear progreso de carga
                task_id = result.get('task_id')
                if task_id:
                    success = monitor_upload_progress(base_url, task_id)
                    if success:
                        successful_files += 1
                    else:
                        failed_files.append((file_name, "Fallo en el procesamiento"))

            except requests.exceptions.RequestException as e:
                print(f"\n⚠️ Error al subir documento {file_name}: {e}")
                failed_files.append((file_name, str(e)))

            # Cerrar handlers de archivos
            for file_entry in files_to_upload:
                _, file_obj = file_entry
                file_obj[1].close()

        except Exception as e:
            print(f"\n⚠️ Error procesando archivo {file_name}: {e}")
            failed_files.append((file_name, str(e)))
        
        # Actualizar progreso
        file_progress.update(1)

    file_progress.close()
    
    # Resumen final
    print(f"\n📊 Resumen de procesamiento:")
    print(f"  ✅ Archivos procesados correctamente: {successful_files}/{len(files)}")
    
    if failed_files:
        print(f"  ⚠️ Archivos con advertencias: {len(failed_files)}")
        for name, error in failed_files[:5]:  # Mostrar los primeros 5 errores
            print(f"     - {name}: {error}")
        
        if len(failed_files) > 5:
            print(f"     ... y {len(failed_files) - 5} más")
            
    file_progress.close()
    return successful_files == len(files)


def monitor_upload_progress(base_url, task_id):
    """
    Monitorea el progreso de un proceso de carga de documentos.
    
    Returns:
        bool: True si la carga se completó con éxito, False en caso contrario
    """
    progress_bar = tqdm(total=100, desc="Procesando", unit="%", leave=False)
    last_percentage = 0
    success = False

    while True:
        try:
            # Aumentar el timeout para archivos grandes
            response = requests.get(f"{base_url}/api/documents/progress/{task_id}", timeout=120)
            response.raise_for_status()

            data = response.json()
            status = data.get('status')
            percentage = data.get('percentage', 0)

            # Actualizar barra de progreso
            progress_diff = percentage - last_percentage
            if progress_diff > 0:
                progress_bar.update(progress_diff)
                last_percentage = percentage

            if status == "completed":
                progress_bar.update(100 - last_percentage)  # Asegurar que lleguemos al 100%
                progress_bar.close()
                print(f"✅ Carga completada: {data.get('message', 'Éxito')}")
                success = True
                break
                
            elif status == "error":
                progress_bar.close()
                # Analizar el mensaje de error para ser más descriptivo
                error_message = data.get('message', 'Error desconocido')
                
                # Lista de campos que podrían faltar según los errores comunes
                missing_fields = ["sheet_index", "width", "height", "total_pages", "total_sheets", "sheet_name", "page_number", "source", "file_type"]
                
                # Verificar si es un error de campo faltante en la colección
                if "DataNotMatchException" in error_message:
                    missing_field_found = False
                    for field in missing_fields:
                        if f"missed an field `{field}`" in error_message:
                            print(f"⚠️ Error de campo: Milvus requiere el campo '{field}'")
                            missing_field_found = True
                            break
                    
                    if not missing_field_found:
                        print(f"❌ Error en la carga: {error_message}")
                else:
                    print(f"❌ Error en la carga: {error_message}")
                break
                
            # Verificar timeout - evitar esperar indefinidamente
            elif progress_bar.n == last_percentage and progress_bar.n > 0:
                if progress_bar._time_elapsed() > 120:  # 2 minutos sin cambios
                    progress_bar.close()
                    print("⚠️ Tiempo de espera excedido sin progreso")
                    break

            time.sleep(2)  # Consultar cada 2 segundos

        except requests.exceptions.RequestException as e:
            progress_bar.close()
            print(f"⚠️ Error al monitorear el progreso: {e}")
            break
            
    return success


def get_document_pages(file_path):
    """
    Determina el número de páginas de un documento según su tipo.
    Replica la lógica de utils/loaders.py de ambos proyectos combinando sus fortalezas
    """
    ext = Path(file_path).suffix.lower().lstrip('.')
    
    try:
        # PDF - Usar PyMuPDF para contar páginas
        if ext == 'pdf':
            try:
                import fitz  # PyMuPDF se importa como fitz
                doc = fitz.open(file_path)
                return len(doc)
            except Exception as e:
                print(f"Error al procesar PDF {file_path}: {e}")
                return 1
            
        # Word - Contar saltos de página en el texto
        elif ext in ['doc', 'docx']:
            try:
                import docx2txt
                # Extrae el texto completo del archivo Word
                text = docx2txt.process(file_path)
                
                # Limpia y agrupa párrafos para un mejor procesamiento
                text = clean_extra_whitespace(text)
                text = group_broken_paragraphs(text)
                
                # Dividir por saltos de página (\f) como en loaders.py
                pages = text.split('\f')
                
                # Si no hay saltos de página explícitos, intentamos encontrar 
                # límites "naturales" o simplemente lo tratamos como una página
                if len(pages) <= 1:
                    # En caso de documentos largos sin marcadores de página explícitos,
                    # podríamos intentar dividir por párrafos o secciones, pero
                    # para simplicidad lo dejamos como una página en esta versión
                    pages = [text]
                
                num_pages = max(1, len(pages))  # Al menos 1 página
                print(f"Word - Número de páginas detectadas: {num_pages}")
                print(f"Word - Usando dimensiones predeterminadas: 612x864")
                return num_pages
            except Exception as e:
                print(f"Error al procesar Word {file_path}: {e}")
                return 1
            
        # Excel - Contar hojas como páginas
        elif ext in ['xls', 'xlsx']:
            try:
                import openpyxl
                # Cargar el workbook con data_only=True para obtener valores en lugar de fórmulas
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                num_sheets = len(wb.sheetnames)
                
                # Verificar que hay al menos una hoja
                if num_sheets == 0:
                    print(f"Advertencia: El archivo Excel {file_path} no tiene hojas")
                    return 1
                
                print(f"Excel - Número de hojas detectadas: {num_sheets}")
                print(f"Excel - Usando dimensiones predeterminadas: 612x864")
                return num_sheets
            except Exception as e:
                print(f"Error al procesar Excel {file_path}: {e}")
                return 1
            
        # Tipo no soportado o desconocido
        else:
            print(f"Tipo de archivo no soportado específicamente: {ext}, usando 1 página por defecto")
            return 1  # Por defecto, asumimos 1 página
            
    except Exception as e:
        print(f"Error general al determinar páginas de {file_path}: {e}")
        return 1  # En caso de error, asumimos 1 página
        
# Funciones auxiliares para procesar texto - similares a las usadas en ambos proyectos
def clean_extra_whitespace(text):
    """
    Limpia espacios en blanco extra del texto.
    Convierte múltiples espacios en uno solo.
    """
    # Reemplazar todos los caracteres de espacio con un solo espacio
    return " ".join(text.split())

def group_broken_paragraphs(text):
    """
    Agrupa párrafos rotos por saltos de línea.
    """
    # Reemplazar saltos de línea con espacios para mejorar la lectura
    return text.replace("\n", " ").replace("\r", " ")

def get_sheet_index(file_path):
    """
    Determina el índice de la hoja para documentos Excel.
    Para otros tipos de documentos, devuelve 0.
    """
    ext = Path(file_path).suffix.lower().lstrip('.')
    
    try:
        # Solo para Excel calculamos el sheet_index
        if ext in ['xls', 'xlsx']:
            try:
                # Para la carga de documentos, usamos la primera hoja (índice 0)
                # ya que cada hoja se procesa como un documento separado
                return 0
            except Exception as e:
                print(f"Error al determinar sheet_index para Excel {file_path}: {e}")
                return 0
        else:
            # Para otros tipos de documentos, sheet_index es 0
            return 0
            
    except Exception as e:
        print(f"Error general al determinar sheet_index de {file_path}: {e}")
        return 0  # Valor por defecto en caso de error

def get_sheet_name(file_path):
    """
    Determina el nombre de la hoja para documentos Excel.
    Para otros tipos de documentos, devuelve una cadena vacía.
    Implementa el enfoque utilizado en local-adv-rag combinado con basic-rag-new.
    """
    ext = Path(file_path).suffix.lower().lstrip('.')
    
    try:
        # Solo para Excel obtenemos el nombre de la hoja
        if ext in ['xls', 'xlsx']:
            try:
                import openpyxl
                # Cargar con data_only=True como hace basic-rag-new para obtener valores no fórmulas
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                # Verificar que hay hojas disponibles
                if wb.sheetnames:
                    sheet_name = wb.sheetnames[0]  # Nombre de la primera hoja
                    print(f"Excel - Nombre de hoja: {sheet_name}")
                    
                    # Verificar que el nombre de la hoja no está vacío
                    if not sheet_name or sheet_name.strip() == "":
                        sheet_name = f"Sheet_{0}"  # Nombre predeterminado
                        print(f"Excel - Usando nombre de hoja predeterminado: {sheet_name}")
                    
                    return sheet_name
                else:
                    # Si no hay hojas, usar un nombre predeterminado
                    sheet_name = f"Sheet_{0}"
                    print(f"Excel - Sin hojas, usando nombre predeterminado: {sheet_name}")
                    return sheet_name
            except Exception as e:
                print(f"Error al determinar sheet_name para Excel {file_path}: {e}")
                # En caso de error, devolver un nombre genérico
                return f"Sheet_{0}"
        else:
            # Para otros tipos de documentos, sheet_name es vacío
            return ""
            
    except Exception as e:
        print(f"Error general al determinar sheet_name de {file_path}: {e}")
        return f"Sheet_{0}"  # Valor predeterminado en caso de error

def get_document_dimensions(file_path):
    """
    Determina el ancho y alto de un documento según su tipo.
    Para PDF, intenta obtener las dimensiones reales del documento.
    Para otros tipos, usa valores estándar.
    Implementa el enfoque combinado de ambos proyectos.
    
    Returns:
        tuple: (width, height) - dimensiones del documento
    """
    ext = Path(file_path).suffix.lower().lstrip('.')
    
    # Valores por defecto para todos los tipos de documentos - A4 a 72 DPI
    # Estos valores son comunes en ambos proyectos
    DEFAULT_WIDTH = 612   # 8.5 pulgadas a 72 DPI
    DEFAULT_HEIGHT = 864  # 12 pulgadas a 72 DPI (mayor que A4 para asegurar espacio)
    
    try:
        # PDF - Obtener dimensiones reales de la primera página
        if ext == 'pdf':
            try:
                import fitz  # PyMuPDF se importa como fitz
                doc = fitz.open(file_path)
                if len(doc) > 0:
                    page = doc[0]  # Primera página
                    width = page.rect.width
                    height = page.rect.height
                    print(f"PDF - Dimensiones calculadas: {width}x{height}")
                    
                    # Verificar que las dimensiones son valores razonables
                    if width <= 0 or height <= 0:
                        print(f"PDF con dimensiones inválidas, usando dimensiones predeterminadas")
                        return DEFAULT_WIDTH, DEFAULT_HEIGHT
                        
                    return float(width), float(height)  # Asegurar que sean float
                else:
                    print(f"PDF sin páginas, usando dimensiones predeterminadas")
                    return DEFAULT_WIDTH, DEFAULT_HEIGHT
            except Exception as e:
                print(f"Error al procesar PDF para dimensiones {file_path}: {e}")
                return DEFAULT_WIDTH, DEFAULT_HEIGHT
                
        # Word - usar dimensiones estándar (A4)
        elif ext in ['doc', 'docx']:
            # En basic-rag-new y local-adv-rag ambos usan dimensiones estándar para Word
            print(f"Word - Usando dimensiones predeterminadas: {DEFAULT_WIDTH}x{DEFAULT_HEIGHT}")
            return DEFAULT_WIDTH, DEFAULT_HEIGHT
        
        # Excel - usar dimensiones estándar
        elif ext in ['xls', 'xlsx']:
            # En basic-rag-new y local-adv-rag ambos usan dimensiones estándar para Excel
            print(f"Excel - Usando dimensiones predeterminadas: {DEFAULT_WIDTH}x{DEFAULT_HEIGHT}")
            return DEFAULT_WIDTH, DEFAULT_HEIGHT
            
        # Otros tipos de archivos
        else:
            print(f"Tipo no específico: {ext} - Usando dimensiones predeterminadas: {DEFAULT_WIDTH}x{DEFAULT_HEIGHT}")
            return DEFAULT_WIDTH, DEFAULT_HEIGHT
            
    except Exception as e:
        print(f"Error general al determinar dimensiones de {file_path}: {e}")
        return DEFAULT_WIDTH, DEFAULT_HEIGHT  # Valores por defecto en caso de error

def get_document_width(file_path):
    """
    Determina el ancho de un documento según su tipo.
    Función de compatibilidad que utiliza get_document_dimensions.
    
    Returns:
        float: ancho del documento
    """
    width, _ = get_document_dimensions(file_path)
    return width


def check_api_schema(base_url):
    """
    Realiza una verificación directa al API para obtener información sobre el schema.
    Esta función es útil para depurar problemas con los campos de metadatos.
    """
    print("\n🔍 Verificando API y esquema de metadatos...\n")
    
    # 1. Verificar conexión básica al API
    try:
        response = requests.get(f"{base_url}/docs", timeout=10)
        if response.status_code == 200:
            print("✅ Conexión al API verificada (documentación disponible)")
        else:
            print(f"⚠️ La documentación del API no está disponible. Status: {response.status_code}")
    except Exception as e:
        print(f"❌ Error de conexión a {base_url}/docs: {e}")
        
    # 2. Verificar si hay colecciones existentes
    try:
        response = requests.get(f"{base_url}/api/documents/collections", timeout=10)
        if response.status_code == 200:
            collections = response.json()
            print(f"✅ API de colecciones accesible. Colecciones encontradas: {len(collections)}")
            
            # Mostrar algunas colecciones para referencia
            if collections:
                print("📚 Primeras 3 colecciones en el sistema:")
                for idx, coll in enumerate(collections[:3]):
                    print(f"   {idx+1}. {coll.get('name', 'Sin nombre')} - Documentos: {coll.get('count', 'N/A')}")
        else:
            print(f"⚠️ No se pudo acceder a la lista de colecciones. Status: {response.status_code}")
    except Exception as e:
        print(f"❌ Error accediendo a colecciones: {e}")
        
    print("\n🔍 Verificación de API y esquema finalizada\n")

def main():
    """Función principal."""
    parser = argparse.ArgumentParser(description='Carga masiva de documentos para RAG desde carpetas de idioma.')
    parser.add_argument('--url', default='http://localhost:8000', help='URL base de la API de RAG')
    parser.add_argument('--dir', help='Directorio base que contiene carpetas "de" y "en"')
    parser.add_argument('--collection', default='documents', help='Nombre base de la colección')
    parser.add_argument('--test', action='store_true', help='Modo de prueba - agrega un sufijo de timestamp a la colección')
    parser.add_argument('--check-schema', action='store_true', help='Verificar esquema API sin cargar documentos')

    args = parser.parse_args()
    
    # Si está activada la opción de verificar schema, solo hacemos eso
    if args.check_schema:
        check_api_schema(args.url)
        return
    
    # Para cargar documentos, el directorio es obligatorio
    if not args.dir:
        parser.error("El argumento --dir es obligatorio para cargar documentos")
        
    # Si estamos en modo de prueba, agregar el timestamp a la colección
    collection_name = args.collection
    if args.test:
        import time
        timestamp = int(time.time())
        collection_name = f"{args.collection}_test_{timestamp}"
        print(f"Modo de prueba activado. Usando colección: {collection_name}")

    upload_documents(args.url, args.dir, collection_name)

if __name__ == "__main__":
    main()