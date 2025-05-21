import os
import re
import pymupdf
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from langchain.docstore.document import Document
from langchain_community.document_loaders import WebBaseLoader
from utils.files import get_file_type
import docx2txt
import re
from typing import List
from langchain.schema import Document
import openpyxl

os.getenv("USER_AGENT")

def load_documents(folder_path):
    documents = []
    for file in os.listdir(folder_path):
        if not file.startswith('~'):
            file_path = os.path.join(folder_path, file)
            
            # Read file as bytes
            with open(file_path, 'rb') as f:
                file_bytes = f.read()
            
            # Use get_file_type from files.py
            file_extension, mime_type = get_file_type(file_bytes)
            
            if file_extension:
                if file_extension == FileType.PDF:
                    documents.extend(load_pdf(file_path, file))
                elif file_extension in [FileType.WORD, FileType.WORDX]:
                    documents.extend(load_docx(file_path, file))
                elif file_extension in [FileType.EXCEL, FileType.EXCELX]:
                    documents.extend(load_xlsx(file_path, file))

    return documents


def load_pdf(file, filename, page_overlap=256):
    documents = []
    try:
        # Open the PDF file
        doc = pymupdf.open(file, filetype="pdf")

        # Extract text from each page
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text = page.get_text("text")
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)

            # Obtener el overlap de la página anterior.
            if page_num > 0:
                prev_page = doc.load_page(page_num - 1)
                prev_page_text = prev_page.get_text("text")
                prev_page_text = clean_extra_whitespace(prev_page_text)
                prev_page_text = group_broken_paragraphs(prev_page_text)
                text = prev_page_text[-page_overlap:] + " " + text

            # Obtener el overlap de la página siguiente.
            if page_num < len(doc) - 1:
                next_page = doc.load_page(page_num + 1)
                next_page_text = next_page.get_text("text")
                next_page_text = clean_extra_whitespace(next_page_text)
                next_page_text = group_broken_paragraphs(next_page_text)
                text += " " + next_page_text[:page_overlap]

            metadata = {
                "source": filename,
                "file_type": FileType.PDF,
                "page_number": page_num + 1,
                "sheet_name": "",
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(e)
        raise

    return documents


def load_webpage(url):
    documents = []
    try:
        loader = WebBaseLoader(url)
        docs = loader.load()
        for doc in docs:
            text = doc.page_content
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)

            metadata = {
                "source": url,
                "file_type": FileType.WEBPAGE,
                "page_number": -1,
                "sheet_name": "",
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(f"An error occurred while loading {e}: {e}")
        raise

    return documents


def load_word(file, filename):
    documents = []
    try:
        parser = LlamaParse(result_type="markdown")
        docs = parser.aload_data(file, extra_info={"file_name": filename})
        for page_num, doc in enumerate(docs):
            text = doc.text
            metadata = {
                "source": filename,
                "file_type": FileType.WORD,
                "page_number": page_num + 1,
                "sheet_name": "",
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(e)
        raise

    return documents


def split_into_pages(text: str) -> List[str]:
    """
    Divide el texto en páginas basándose en saltos de página.
    """
    return text.split('\f')

def load_docx(file_path: str, filename: str) -> List[Document]:
    """
    Carga texto de un archivo DOCX, lo divide en páginas y crea objetos Document con overlap.
    """
    # Extraer texto del archivo DOCX
    text = docx2txt.process(file_path)
    
    # Limpiar y agrupar párrafos
    text = clean_extra_whitespace(text)
    text = group_broken_paragraphs(text)
    
    # Dividir en páginas
    pages = split_into_pages(text)
    
    documents = []
    overlap_words = 8

    for i, page_content in enumerate(pages):
        # Preparar el contenido con overlap
        content_parts = []
        
        # Agregar overlap antes
        if i > 0:
            prev_page_words = pages[i-1].split()
            content_parts.append(" ".join(prev_page_words[-overlap_words:]))
        
        # Agregar contenido de la página actual
        content_parts.append(page_content)
        
        # Agregar overlap después
        if i < len(pages) - 1:
            next_page_words = pages[i+1].split()
            content_parts.append(" ".join(next_page_words[:overlap_words]))
        
        # Crear objeto Document
        doc = Document(
            page_content=" ".join(content_parts),
            metadata = {
                "source": filename,
                "file_type": FileType.WORD,
                "page_number": i + 1,
                "sheet_name": "",
            }
        )
        documents.append(doc)
      
    return documents


def load_excel(file, filename):
    documents = []
    try:
        parser = LlamaParse(result_type="markdown")
        docs = parser.aload_data(file, extra_info={"file_name": filename})
        for doc in docs:
            text = doc.text
            metadata = {
                "source": filename,
                "file_type": FileType.EXCEL,
                "page_number": -1,
                "sheet_name": extract_sheet_name(text),
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(e)
        raise

    return documents



def load_xlsx(file_path, filename):
    """
    Loads text from an XLSX file, including sheet names in metadata.
    """
    wb = openpyxl.load_workbook(file_path)
    documents = []
    for sheet_num, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        text = ""
        for row in ws.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        text = clean_extra_whitespace(text)
        text = group_broken_paragraphs(text)
        if text:  # Only add if the text is not empty
            document = Document(
                page_content=text,
                metadata = {
                    "source": filename,
                    "file_type": FileType.EXCEL,
                    "page_number": -1,
                    "sheet_name": sheet_name,
                }
            )
            documents.append(document)

    return documents


def clean_extra_whitespace(text):
    """
    Cleans extra whitespace from the provided text.

    Parameters:
    - text: A string representing the text to be cleaned.

    Returns:
    - A string with extra whitespace removed.
    """
    return " ".join(text.split())


def group_broken_paragraphs(text):
    """
    Groups broken paragraphs in the provided text.

    Parameters:
    - text: A string representing the text to be processed.

    Returns:
    - A string with broken paragraphs grouped.
    """
    return text.replace("\n", " ").replace("\r", " ")


def extract_sheet_name(text):
    match = re.match(r"^# (.*?)\n", text)
    if match:
        return match.group(1)
    return None


# Función para cargar una única página de Confluence utilizando autenticación
def load_confluence(url, username, password):
    """
    Carga una única página de Confluence utilizando autenticación,
    limpia el contenido y lo devuelve como una lista con un objeto Document.
    
    Parámetros:
    - url: URL de la página en Confluence.
    - username: Nombre de usuario para la autenticación.
    - password: Contraseña para la autenticación.
    
    Returns:
    - Una lista con un objeto Document.
    """
    documents = []
    session = requests.Session()
    session.auth = (username, password)
    try:
        response = session.get(url)
        response.raise_for_status()
        html = response.text
        # Utilizar BeautifulSoup para extraer el contenido relevante
        soup = BeautifulSoup(html, "html.parser")
        # Se asume que el contenido principal se encuentra en la etiqueta <body>
        content_elem = soup.find("body")
        text = content_elem.get_text(separator=" ", strip=True) if content_elem else html
        text = clean_extra_whitespace(text)
        text = group_broken_paragraphs(text)
        
        metadata = {
            "source": url,
            "file_type": "confluence",
            "page_number": -1,
            "sheet_name": "",
        }
        document = Document(page_content=text, metadata=metadata)
        documents.append(document)
    except Exception as e:
        print(f"Ocurrió un error al cargar la página de Confluence: {e}")
        raise
    return documents


# Nueva función para cargar la estructura de páginas (árbol) de Confluence de forma genérica
def load_confluence_tree(base_url, username, password):
    """
    Dada la URL base de un espacio de Confluence (por ejemplo, https://wiki.uni-graz.at/display/uniK/),
    extrae de la página principal todos los enlaces hacia las páginas hijas pertenecientes al mismo espacio
    y carga el contenido de cada una de ellas.
    
    Parámetros:
    - base_url: URL base del espacio en Confluence.
    - username: Nombre de usuario para la autenticación.
    - password: Contraseña para la autenticación.
    
    Returns:
    - Una lista de objetos Document correspondientes a cada página extraída.
    """
    import requests
    from urllib.parse import urljoin, urlparse

    documents = []
    session = requests.Session()
    session.auth = (username, password)
    try:
        response = session.get(base_url)
        response.raise_for_status()
        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        body = soup.find("body")
        if not body:
            return documents

        # Obtener el patrón base a partir de la URL base, por ejemplo "display/uniK"
        parsed_url = urlparse(base_url)
        path = parsed_url.path.rstrip("/")  # e.g. "/display/uniK"
        base_pattern = path.lstrip("/")      # e.g. "display/uniK"

        # Extraer todos los enlaces que contengan el patrón base de la URL (para el espacio correspondiente)
        links = body.find_all("a", href=True)
        urls = set()
        for link in links:
            href = link["href"]
            if base_pattern in href:
                full_url = urljoin(base_url, href)
                urls.add(full_url)

        # Cargar cada página encontrada
        for url in urls:
            docs = load_confluence(url, username, password)
            documents.extend(docs)
    except Exception as e:
        print(f"Ocurrió un error al cargar la estructura de Confluence: {e}")
        raise
    return documents


# Tipos de archivo y extensiones permitidas
class FileType:
    PDF = "pdf"
    WORD = "doc"
    WORDX = "docx"
    EXCEL = "xls"
    EXCELX = "xlsx"
    WEBPAGE = "html"
