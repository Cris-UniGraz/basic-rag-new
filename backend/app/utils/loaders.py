import os
import re
import pymupdf
import requests
import asyncio
import traceback
from pathlib import Path
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from typing import List, Dict, Any, Optional, Union, Tuple
from langchain_core.documents import Document
from langchain_community.document_loaders import WebBaseLoader
import docx2txt
import openpyxl
from loguru import logger
from tenacity import retry, stop_after_attempt, wait_exponential
from concurrent.futures import ThreadPoolExecutor
import time

from app.utils.files import get_file_type, calculate_document_size
from app.core.metrics import measure_time, DOCUMENT_PROCESSING_DURATION
from app.core.config import settings


class FileType:
    """Constants for supported file types."""
    PDF = "pdf"
    WORD = "doc"
    WORDX = "docx"
    EXCEL = "xls"
    EXCELX = "xlsx"
    WEBPAGE = "html"
    TXT = "txt"
    CSV = "csv"
    JSON = "json"
    MARKDOWN = "md"


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": "batch"})
def load_documents(
    paths: Union[str, Path, List[Union[str, Path]]],
    recursive: bool = False,
    file_types: Optional[List[str]] = None,
    parallel: bool = True,
    max_workers: int = 5
) -> List[Document]:
    """
    Load documents from a directory or a list of file paths with optimized parallel processing.
    
    Args:
        paths: Path to the directory containing documents or a list of file paths
        recursive: Whether to search subdirectories (only used if paths is a directory)
        file_types: List of file extensions to include (None for all supported types)
        parallel: Whether to use parallel processing
        max_workers: Maximum number of worker threads for parallel processing
        
    Returns:
        List of Document objects
    """
    documents = []
    file_paths = []
    
    # Get supported file types if not specified
    if file_types is None:
        file_types = [
            FileType.PDF, FileType.WORD, FileType.WORDX, 
            FileType.EXCEL, FileType.EXCELX, FileType.TXT,
            FileType.CSV, FileType.MARKDOWN, FileType.JSON
        ]
    
    # Convert file types to lowercase
    file_types = [ft.lower() for ft in file_types]
    
    # Handle different input types
    if isinstance(paths, list):
        # Input is already a list of file paths
        file_paths = [Path(path) for path in paths]
        logger.info(f"Processing {len(file_paths)} provided files")
    else:
        # Input is a directory path
        folder_path = Path(paths)
        logger.info(f"Searching for documents in {folder_path}")
        start_time = time.time()
        
        # Recursively find files if requested
        if recursive:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if not file.startswith('~') and not file.startswith('.'):
                        file_path = Path(root) / file
                        file_paths.append(file_path)
        else:
            # Just look in the specified directory
            for file in os.listdir(folder_path):
                if not file.startswith('~') and not file.startswith('.'):
                    file_path = folder_path / file
                    if file_path.is_file():
                        file_paths.append(file_path)
        
        search_time = time.time() - start_time
        logger.info(f"Found {len(file_paths)} files in {search_time:.2f}s")
    
    # Function to process a single file
    def process_file(file_path: Path) -> List[Document]:
        try:
            # Check if file extension is in the target list
            suffix = file_path.suffix.lower().lstrip('.')
            filename = file_path.name
            
            # If extension is not in our list, try to determine file type
            if suffix not in file_types:
                with open(file_path, 'rb') as f:
                    file_bytes = f.read(1024)  # Read just enough to determine file type
                detected_ext, _ = get_file_type(file_bytes)
                
                if detected_ext and detected_ext.lower() in file_types:
                    suffix = detected_ext.lower()
                else:
                    # Skip this file as it's not a target type
                    return []
            
            # Process the file based on its type
            if suffix == FileType.PDF:
                return load_pdf(str(file_path), filename)
            elif suffix in [FileType.WORD, FileType.WORDX]:
                return load_docx(str(file_path), filename)
            elif suffix in [FileType.EXCEL, FileType.EXCELX]:
                return load_xlsx(str(file_path), filename)
            elif suffix == FileType.TXT:
                return load_text(str(file_path), filename)
            elif suffix == FileType.CSV:
                return load_csv(str(file_path), filename)
            elif suffix == FileType.MARKDOWN:
                return load_markdown(str(file_path), filename)
            elif suffix == FileType.JSON:
                return load_json(str(file_path), filename)
            else:
                logger.warning(f"Unsupported file type: {suffix} for {file_path}")
                return []
                
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            logger.debug(traceback.format_exc())
            return []
    
    # Process files either in parallel or sequentially
    if parallel and len(file_paths) > 1:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = list(executor.map(process_file, file_paths))
        
        # Flatten results
        for result in results:
            if result:
                documents.extend(result)
    else:
        for file_path in file_paths:
            try:
                docs = process_file(file_path)
                documents.extend(docs)
            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
    
    return documents


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.PDF})
def load_pdf(
    file: Union[str, Path], 
    filename: str, 
    page_overlap: int = None
) -> List[Document]:
    """
    Load and process a PDF file with text extraction and overlap handling.
    
    Args:
        file: Path to the PDF file
        filename: Name of the file (for metadata)
        page_overlap: Number of characters to overlap between pages
        
    Returns:
        List of Document objects, one per page
    """
    if page_overlap is None:
        page_overlap = settings.PAGE_OVERLAP
        
    documents = []
    
    try:
        # Open the PDF file
        doc = pymupdf.open(file, filetype="pdf")
        total_pages = len(doc)
        
        logger.debug(f"Processing PDF: {filename} ({total_pages} pages)")
        
        # Extract text from each page
        for page_num in range(total_pages):
            page = doc.load_page(page_num)
            text = page.get_text("text")
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)
            
            # Get overlap from previous page
            if page_num > 0:
                prev_page = doc.load_page(page_num - 1)
                prev_page_text = prev_page.get_text("text")
                prev_page_text = clean_extra_whitespace(prev_page_text)
                prev_page_text = group_broken_paragraphs(prev_page_text)
                
                # Only add overlap if there's enough text
                if len(prev_page_text) > page_overlap:
                    text = prev_page_text[-page_overlap:] + " " + text
            
            # Get overlap from next page
            if page_num < total_pages - 1:
                next_page = doc.load_page(page_num + 1)
                next_page_text = next_page.get_text("text")
                next_page_text = clean_extra_whitespace(next_page_text)
                next_page_text = group_broken_paragraphs(next_page_text)
                
                # Only add overlap if there's enough text
                if len(next_page_text) > page_overlap:
                    text += " " + next_page_text[:page_overlap]
            
            # Create metadata
            metadata = {
                "source": filename,
                "file_path": str(file),
                "file_type": FileType.PDF,
                "page_number": page_num + 1,
                "total_pages": total_pages,
                "sheet_name": "",
            }
            
            # Extract page dimensions for additional metadata
            try:
                rect = page.rect
                metadata["width"] = rect.width
                metadata["height"] = rect.height
            except:
                pass
            
            # Create document
            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
            
        logger.debug(f"Extracted {len(documents)} pages from PDF: {filename}")
        
    except Exception as e:
        logger.error(f"Error processing PDF {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise
    
    return documents


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.WEBPAGE})
def load_webpage(
    url: str, 
    user_agent: Optional[str] = None
) -> List[Document]:
    """
    Load and process a webpage.
    
    Args:
        url: URL of the webpage
        user_agent: User agent string for the request
        
    Returns:
        List containing a single Document object with the webpage content
    """
    documents = []
    
    try:
        # Configure loader with user agent if provided
        headers = {}
        if user_agent or settings.USER_AGENT:
            headers["User-Agent"] = user_agent or settings.USER_AGENT
        
        loader = WebBaseLoader(url, header_template=headers)
        docs = loader.load()
        
        # Process each document
        for doc in docs:
            text = doc.page_content
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)
            
            # Create metadata
            metadata = {
                "source": url,
                "file_type": FileType.WEBPAGE,
                "page_number": -1,
                "sheet_name": "",
                "title": extract_webpage_title(text),
            }
            
            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
            
        logger.debug(f"Loaded webpage: {url}")
        
    except Exception as e:
        logger.error(f"Error loading webpage {url}: {e}")
        logger.debug(traceback.format_exc())
        raise
    
    return documents


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.WORDX})
def load_docx(
    file_path: Union[str, Path], 
    filename: str, 
    overlap_words: int = 8
) -> List[Document]:
    """
    Load and process a DOCX file.
    
    Args:
        file_path: Path to the DOCX file
        filename: Name of the file (for metadata)
        overlap_words: Number of words to overlap between pages
        
    Returns:
        List of Document objects
    """
    try:
        # Extract text from DOCX file
        text = docx2txt.process(file_path)
        
        # Clean and group paragraphs
        text = clean_extra_whitespace(text)
        text = group_broken_paragraphs(text)
        
        # Split into pages
        pages = split_into_pages(text)
        
        documents = []
        
        # Process each page
        for i, page_content in enumerate(pages):
            # Prepare content with overlap
            content_parts = []
            
            # Add overlap from previous page
            if i > 0:
                prev_page_words = pages[i-1].split()
                if len(prev_page_words) >= overlap_words:
                    content_parts.append(" ".join(prev_page_words[-overlap_words:]))
            
            # Add current page content
            content_parts.append(page_content)
            
            # Add overlap from next page
            if i < len(pages) - 1:
                next_page_words = pages[i+1].split()
                if len(next_page_words) >= overlap_words:
                    content_parts.append(" ".join(next_page_words[:overlap_words]))
            
            # Create document
            doc = Document(
                page_content=" ".join(content_parts),
                metadata={
                    "source": filename,
                    "file_path": str(file_path),
                    "file_type": FileType.WORDX,
                    "page_number": i + 1,
                    "total_pages": len(pages),
                    "sheet_name": "",
                }
            )
            documents.append(doc)
        
        logger.debug(f"Extracted {len(documents)} pages from DOCX: {filename}")
        return documents
        
    except Exception as e:
        logger.error(f"Error processing DOCX {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.EXCELX})
def load_xlsx(
    file_path: Union[str, Path], 
    filename: str
) -> List[Document]:
    """
    Load and process an XLSX file, creating one document per sheet.
    
    Args:
        file_path: Path to the XLSX file
        filename: Name of the file (for metadata)
        
    Returns:
        List of Document objects, one per sheet
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        documents = []
        
        # Process each sheet
        for sheet_num, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            text = ""
            
            # Extract cell values
            for row in ws.iter_rows(values_only=True):
                # Skip empty rows
                if not any(cell is not None for cell in row):
                    continue
                
                # Join cell values
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text += row_text + "\n"
            
            # Clean text
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)
            
            # Skip empty sheets
            if not text.strip():
                continue
            
            # Create document
            document = Document(
                page_content=text,
                metadata={
                    "source": filename,
                    "file_path": str(file_path),
                    "file_type": FileType.EXCELX,
                    "page_number": -1,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_num,
                    "total_sheets": len(wb.sheetnames),
                }
            )
            documents.append(document)
        
        logger.debug(f"Extracted {len(documents)} sheets from XLSX: {filename}")
        return documents
        
    except Exception as e:
        logger.error(f"Error processing XLSX {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.TXT})
def load_text(
    file_path: Union[str, Path], 
    filename: str,
    encoding: str = 'utf-8',
    chunk_size: int = 4000,
    chunk_overlap: int = 200
) -> List[Document]:
    """
    Load and process a text file.
    
    Args:
        file_path: Path to the text file
        filename: Name of the file (for metadata)
        encoding: File encoding
        chunk_size: Size of text chunks to create
        chunk_overlap: Overlap between chunks
        
    Returns:
        List of Document objects
    """
    try:
        # Try to read with specified encoding, fallback to latin-1
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                text = f.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as f:
                text = f.read()
        
        # Clean text
        text = clean_extra_whitespace(text)
        
        # Split into chunks if text is large
        chunks = []
        if len(text) > chunk_size:
            current_chunk = ""
            words = text.split()
            
            for word in words:
                # If adding this word would exceed chunk size, finalize the chunk
                if len(current_chunk) + len(word) + 1 > chunk_size and current_chunk:
                    chunks.append(current_chunk)
                    # Start new chunk with overlap
                    overlap_start = max(0, len(current_chunk) - chunk_overlap)
                    current_chunk = current_chunk[overlap_start:] + " " + word
                else:
                    if current_chunk:
                        current_chunk += " " + word
                    else:
                        current_chunk = word
            
            # Add the last chunk if it has content
            if current_chunk:
                chunks.append(current_chunk)
        else:
            chunks = [text]
        
        # Create documents
        documents = []
        for i, chunk in enumerate(chunks):
            document = Document(
                page_content=chunk,
                metadata={
                    "source": filename,
                    "file_path": str(file_path),
                    "file_type": FileType.TXT,
                    "chunk_number": i + 1,
                    "total_chunks": len(chunks),
                    "sheet_name": "",
                }
            )
            documents.append(document)
        
        logger.debug(f"Extracted {len(documents)} chunks from text file: {filename}")
        return documents
        
    except Exception as e:
        logger.error(f"Error processing text file {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.CSV})
def load_csv(
    file_path: Union[str, Path], 
    filename: str,
    encoding: str = 'utf-8',
) -> List[Document]:
    """
    Load and process a CSV file.
    
    Args:
        file_path: Path to the CSV file
        filename: Name of the file (for metadata)
        encoding: File encoding
        
    Returns:
        List containing a single Document object with the CSV content
    """
    try:
        # Try to read with specified encoding, fallback to latin-1
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                text = f.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as f:
                text = f.read()
        
        # Clean text (preserve newlines for CSV structure)
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Create document
        document = Document(
            page_content=text,
            metadata={
                "source": filename,
                "file_path": str(file_path),
                "file_type": FileType.CSV,
                "page_number": 1,
                "total_pages": 1,
                "sheet_name": "",
            }
        )
        
        logger.debug(f"Loaded CSV file: {filename}")
        return [document]
        
    except Exception as e:
        logger.error(f"Error processing CSV file {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.MARKDOWN})
def load_markdown(
    file_path: Union[str, Path], 
    filename: str,
    encoding: str = 'utf-8',
) -> List[Document]:
    """
    Load and process a Markdown file.
    
    Args:
        file_path: Path to the Markdown file
        filename: Name of the file (for metadata)
        encoding: File encoding
        
    Returns:
        List containing a single Document object with the Markdown content
    """
    try:
        # Try to read with specified encoding, fallback to latin-1
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                text = f.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as f:
                text = f.read()
        
        # Extract title if present (# Title)
        title = ""
        title_match = re.match(r'^#\s+(.+)$', text.split('\n')[0])
        if title_match:
            title = title_match.group(1)
        
        # Create document
        document = Document(
            page_content=text,
            metadata={
                "source": filename,
                "file_path": str(file_path),
                "file_type": FileType.MARKDOWN,
                "page_number": 1,
                "total_pages": 1,
                "sheet_name": "",
                "title": title,
            }
        )
        
        logger.debug(f"Loaded Markdown file: {filename}")
        return [document]
        
    except Exception as e:
        logger.error(f"Error processing Markdown file {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise


@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": FileType.JSON})
def load_json(
    file_path: Union[str, Path], 
    filename: str,
    encoding: str = 'utf-8',
) -> List[Document]:
    """
    Load and process a JSON file.
    
    Args:
        file_path: Path to the JSON file
        filename: Name of the file (for metadata)
        encoding: File encoding
        
    Returns:
        List containing a single Document object with the JSON content
    """
    try:
        # Try to read with specified encoding, fallback to latin-1
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                text = f.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as f:
                text = f.read()
        
        # Create document
        document = Document(
            page_content=text,
            metadata={
                "source": filename,
                "file_path": str(file_path),
                "file_type": FileType.JSON,
                "page_number": 1,
                "total_pages": 1,
                "sheet_name": "",
            }
        )
        
        logger.debug(f"Loaded JSON file: {filename}")
        return [document]
        
    except Exception as e:
        logger.error(f"Error processing JSON file {filename}: {e}")
        logger.debug(traceback.format_exc())
        raise


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": "confluence"})
def load_confluence(
    url: str, 
    username: str, 
    password: str
) -> List[Document]:
    """
    Load a Confluence page with authentication.
    
    Args:
        url: URL of the Confluence page
        username: Authentication username
        password: Authentication password
        
    Returns:
        List containing a single Document object with the page content
    """
    documents = []
    session = requests.Session()
    session.auth = (username, password)
    
    try:
        response = session.get(url, timeout=30)
        response.raise_for_status()
        html = response.text
        
        # Parse HTML
        soup = BeautifulSoup(html, "html.parser")
        
        # Extract title
        title_elem = soup.find("title")
        title = title_elem.get_text() if title_elem else ""
        
        # Extract main content
        content_elem = soup.find("div", id="main-content") or soup.find("body")
        text = content_elem.get_text(separator=" ", strip=True) if content_elem else html
        
        # Clean text
        text = clean_extra_whitespace(text)
        text = group_broken_paragraphs(text)
        
        # Create metadata
        metadata = {
            "source": url,
            "file_type": "confluence",
            "page_number": -1,
            "sheet_name": "",
            "title": title,
        }
        
        document = Document(page_content=text, metadata=metadata)
        documents.append(document)
        
        logger.debug(f"Loaded Confluence page: {url}")
        
    except Exception as e:
        logger.error(f"Error loading Confluence page {url}: {e}")
        logger.debug(traceback.format_exc())
        raise
    
    return documents


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
@measure_time(DOCUMENT_PROCESSING_DURATION, {"file_type": "confluence_tree"})
def load_confluence_tree(
    base_url: str, 
    username: str, 
    password: str
) -> List[Document]:
    """
    Load a Confluence space and extract all linked pages.
    
    Args:
        base_url: Base URL of the Confluence space
        username: Authentication username
        password: Authentication password
        
    Returns:
        List of Document objects for all pages in the space
    """
    documents = []
    session = requests.Session()
    session.auth = (username, password)
    
    try:
        # Get the main page
        response = session.get(base_url, timeout=30)
        response.raise_for_status()
        html = response.text
        
        # Parse HTML
        soup = BeautifulSoup(html, "html.parser")
        body = soup.find("body")
        
        if not body:
            logger.warning(f"No body element found in Confluence page: {base_url}")
            return documents
        
        # Get the base pattern to filter links
        parsed_url = urlparse(base_url)
        path = parsed_url.path.rstrip("/")
        base_pattern = path.lstrip("/")
        
        # Find all links in the page
        links = body.find_all("a", href=True)
        urls = set()
        
        # Extract unique URLs that match the pattern
        for link in links:
            href = link["href"]
            if base_pattern in href:
                full_url = urljoin(base_url, href)
                urls.add(full_url)
        
        logger.info(f"Found {len(urls)} links in Confluence space: {base_url}")
        
        # Load each linked page
        with ThreadPoolExecutor(max_workers=5) as executor:
            # Create a list of future tasks
            future_to_url = {
                executor.submit(load_confluence, url, username, password): url 
                for url in urls
            }
            
            # Process results as they complete
            for future in future_to_url:
                url = future_to_url[future]
                try:
                    page_docs = future.result()
                    documents.extend(page_docs)
                except Exception as e:
                    logger.error(f"Error loading Confluence page {url}: {e}")
        
        logger.info(f"Loaded {len(documents)} pages from Confluence space: {base_url}")
        
    except Exception as e:
        logger.error(f"Error loading Confluence tree {base_url}: {e}")
        logger.debug(traceback.format_exc())
        raise
    
    return documents


# Helper functions

def split_into_pages(text: str) -> List[str]:
    """
    Split text into pages based on form feed characters.
    
    Args:
        text: Text to split
        
    Returns:
        List of page texts
    """
    # Split on form feeds
    pages = text.split('\f')
    
    # If there are no form feeds, consider the whole text as one page
    if len(pages) == 1:
        # Try to split on "natural" page boundaries
        pages = split_on_natural_boundaries(text)
    
    # Remove empty pages
    return [page for page in pages if page.strip()]


def split_on_natural_boundaries(text: str, max_length: int = 3000) -> List[str]:
    """
    Split text into pages based on natural boundaries like sections or paragraphs.
    
    Args:
        text: Text to split
        max_length: Maximum length for each page
        
    Returns:
        List of page texts
    """
    # Look for section headings or multiple newlines as natural boundaries
    boundaries = re.finditer(r'(\n\s*#{1,3}\s+[^\n]+\n)|(\n\s*\n\s*\n)', text)
    
    # Get the positions of all boundaries
    positions = [0]  # Start with the beginning of the text
    for match in boundaries:
        positions.append(match.start())
    positions.append(len(text))  # End with the end of the text
    
    # Create pages based on boundaries and max length
    pages = []
    start_idx = 0
    
    for i in range(1, len(positions)):
        # If adding this section would exceed max_length, create a page
        if positions[i] - positions[start_idx] > max_length and start_idx < i - 1:
            pages.append(text[positions[start_idx]:positions[i-1]])
            start_idx = i - 1
    
    # Add the last page
    if start_idx < len(positions) - 1:
        pages.append(text[positions[start_idx]:positions[-1]])
    
    # If we still have large pages, split them by paragraphs
    result = []
    for page in pages:
        if len(page) > max_length:
            result.extend(split_by_paragraphs(page, max_length))
        else:
            result.append(page)
    
    return result


def split_by_paragraphs(text: str, max_length: int = 3000) -> List[str]:
    """
    Split text into chunks based on paragraphs.
    
    Args:
        text: Text to split
        max_length: Maximum length for each chunk
        
    Returns:
        List of text chunks
    """
    paragraphs = re.split(r'\n\s*\n', text)
    chunks = []
    current_chunk = ""
    
    for paragraph in paragraphs:
        # If adding this paragraph would exceed max_length, finalize the chunk
        if len(current_chunk) + len(paragraph) + 2 > max_length and current_chunk:
            chunks.append(current_chunk)
            current_chunk = paragraph
        else:
            if current_chunk:
                current_chunk += "\n\n" + paragraph
            else:
                current_chunk = paragraph
    
    # Add the last chunk
    if current_chunk:
        chunks.append(current_chunk)
    
    return chunks


def clean_extra_whitespace(text: str) -> str:
    """
    Clean extra whitespace from text.
    
    Args:
        text: Text to clean
        
    Returns:
        Cleaned text
    """
    # Replace all whitespace characters with a single space
    return " ".join(text.split())


def group_broken_paragraphs(text: str) -> str:
    """
    Group paragraphs broken by line breaks.
    
    Args:
        text: Text to process
        
    Returns:
        Processed text with line breaks replaced by spaces
    """
    return text.replace("\n", " ").replace("\r", " ")


def extract_webpage_title(html_text: str) -> str:
    """
    Extract title from HTML text.
    
    Args:
        html_text: HTML text
        
    Returns:
        Extracted title or empty string
    """
    # Look for title tags
    title_match = re.search(r'<title>(.*?)</title>', html_text, re.IGNORECASE | re.DOTALL)
    if title_match:
        return title_match.group(1).strip()
    
    # Look for h1 tags
    h1_match = re.search(r'<h1[^>]*>(.*?)</h1>', html_text, re.IGNORECASE | re.DOTALL)
    if h1_match:
        return h1_match.group(1).strip()
    
    # No title found
    return ""


def extract_sheet_name(text: str) -> Optional[str]:
    """
    Extract sheet name from text (usually from Markdown headers).
    
    Args:
        text: Text to process
        
    Returns:
        Extracted sheet name or None
    """
    match = re.match(r'^# (.*?)\n', text)
    if match:
        return match.group(1)
    return None